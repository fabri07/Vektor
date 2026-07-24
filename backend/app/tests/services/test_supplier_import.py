"""Tests del import masivo de proveedores (F7b) — espejo de test_customer_extraction.py.

Cubre:
- ``parse_supplier_records``: planilla → records, reusando el mapeo de columnas de
  ``column_mapping_service`` para ``entity_type="supplier"``.
- ``build_import_preview``: crear/actualizar/needs_review/inválido/duplicado.
- ``apply_import``: upsert idempotente, needs_review/conflicto salteados, sentinela y
  marcas colapsadas nunca tocadas.
"""

from __future__ import annotations

import io
import uuid
from typing import Any

import openpyxl
import pytest

from app.application.services.supplier_extraction_service import parse_supplier_records
from app.application.services.supplier_import_service import apply_import, build_import_preview
from app.persistence.models.supplier import (
    BRAND_COLLAPSED_FLAG_KEY,
    SENTINEL_FLAG_KEY,
    Supplier,
)

_VALID_CUIL = "20-12345678-6"
_VALID_CUIL_2 = "27-23456789-1"


def _xlsx_suppliers(rows: list[list[Any]]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["razon social", "cuil", "email", "telefono"])
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _row(**kw: Any) -> dict[str, Any]:
    return kw


class TestParseSupplierRecords:
    def test_maps_columns_via_column_mapping_service(self) -> None:
        content = _xlsx_suppliers([["Distribuidora Norte", _VALID_CUIL, "a@a.com", "111"]])
        records, warnings = parse_supplier_records(content, "prov.xlsx", "application/octet-stream")
        assert len(records) == 1
        assert records[0]["name"] == "Distribuidora Norte"
        assert records[0]["cuil"] == _VALID_CUIL
        assert records[0]["email"] == "a@a.com"
        assert records[0]["phone"] == "111"
        assert warnings == []

    def test_unrecognized_columns_warn(self) -> None:
        wb_rows = [["columna_random"], ["algo"]]
        wb = openpyxl.Workbook()
        ws = wb.active
        for r in wb_rows:
            ws.append(r)
        buf = io.BytesIO()
        wb.save(buf)
        records, warnings = parse_supplier_records(
            buf.getvalue(), "prov.xlsx", "application/octet-stream"
        )
        assert records == []
        assert any("nombre" in w.lower() for w in warnings)

    def test_no_content_warns(self) -> None:
        wb = openpyxl.Workbook()
        buf = io.BytesIO()
        wb.save(buf)
        records, warnings = parse_supplier_records(
            buf.getvalue(), "prov.xlsx", "application/octet-stream"
        )
        assert records == []
        assert warnings


class TestSupplierImportPreview:
    def test_classifies_create_update_needs_review_duplicate(self) -> None:
        existing = [Supplier(tenant_id=uuid.uuid4(), name="Ya Existe", cuil=_VALID_CUIL)]
        records = [
            _row(name="Nuevo", cuil=_VALID_CUIL_2),  # create
            _row(name="Actualizar", cuil=_VALID_CUIL),  # update (match por cuil)
            _row(name="Sin Dato Fuerte"),  # needs_review
            _row(name="Repe", cuil=_VALID_CUIL_2),  # duplicate_in_file
        ]
        preview = build_import_preview(records, existing)
        assert preview.to_create == 1
        assert preview.to_update == 1
        assert preview.needs_review == 1
        assert preview.invalid == 0
        assert preview.duplicates == 1

    def test_invalid_cuil_check_digit_flagged(self) -> None:
        records = [_row(name="Mal CUIL", cuil="20-12345678-0")]
        preview = build_import_preview(records, [])
        assert preview.invalid == 1

    def test_missing_name_is_invalid(self) -> None:
        records = [_row(cuil=_VALID_CUIL)]
        preview = build_import_preview(records, [])
        assert preview.invalid == 1
        assert preview.needs_review == 0

    def test_email_only_matches_existing(self) -> None:
        existing = [Supplier(tenant_id=uuid.uuid4(), name="Con Email", email="ventas@norte.com")]
        records = [_row(name="Con Email", email="Ventas@Norte.com")]
        preview = build_import_preview(records, existing)
        assert preview.to_update == 1
        assert preview.items[0].existing_id == existing[0].id

    def test_conflict_between_cuil_and_email_is_invalid(self) -> None:
        existing = [
            Supplier(tenant_id=uuid.uuid4(), name="A", cuil=_VALID_CUIL),
            Supplier(tenant_id=uuid.uuid4(), name="B", email="b@b.com"),
        ]
        records = [_row(name="Ambiguo", cuil=_VALID_CUIL, email="b@b.com")]
        preview = build_import_preview(records, existing)
        assert preview.invalid == 1
        assert preview.to_create == 0
        assert preview.to_update == 0


@pytest.mark.asyncio
class TestSupplierApplyImport:
    async def test_upsert_idempotent(self, db_session: Any, sample_tenant: Any) -> None:
        from app.persistence.repositories.supplier_repository import SupplierRepository

        repo = SupplierRepository(db_session)
        tid = sample_tenant.tenant_id
        records = [
            _row(name="Norte", cuil=_VALID_CUIL, payment_method="transferencia"),
            _row(name="Sur", email="sur@sur.com"),
        ]
        first = await apply_import(repo, tid, records)
        assert len(first.created_ids) == 2
        assert first.skipped == 0

        second = await apply_import(repo, tid, records)
        assert len(second.created_ids) == 0
        assert len(second.updated_ids) == 2

        assert await repo.count_active(tid) == 2

    async def test_needs_review_never_created(self, db_session: Any, sample_tenant: Any) -> None:
        from app.persistence.repositories.supplier_repository import SupplierRepository

        repo = SupplierRepository(db_session)
        tid = sample_tenant.tenant_id
        records = [_row(name="Solo Nombre")]
        result = await apply_import(repo, tid, records)
        assert result.created_ids == []
        assert result.skipped == 1
        assert await repo.count_active(tid) == 0

    async def test_conflict_never_created_or_updated(
        self, db_session: Any, sample_tenant: Any
    ) -> None:
        from app.persistence.repositories.supplier_repository import SupplierRepository

        repo = SupplierRepository(db_session)
        tid = sample_tenant.tenant_id
        a = Supplier(tenant_id=tid, name="A", cuil=_VALID_CUIL)
        b = Supplier(tenant_id=tid, name="B", email="b@b.com")
        db_session.add_all([a, b])
        await db_session.commit()

        records = [_row(name="Ambiguo", cuil=_VALID_CUIL, email="b@b.com")]
        result = await apply_import(repo, tid, records)
        assert result.created_ids == []
        assert result.updated_ids == []
        assert result.skipped == 1

    async def test_sentinel_never_touched_by_import(
        self, db_session: Any, sample_tenant: Any
    ) -> None:
        from app.persistence.repositories.supplier_repository import SupplierRepository

        repo = SupplierRepository(db_session)
        tid = sample_tenant.tenant_id
        sentinel = Supplier(
            tenant_id=tid,
            name="No identificado",
            custom_fields={SENTINEL_FLAG_KEY: "true"},
        )
        db_session.add(sentinel)
        await db_session.commit()

        # Import no debería matchear ni actualizar el sentinela aunque comparta CUIL
        # (caso límite: el sentinela típicamente no tiene CUIL, pero igual queda
        # afuera del índice de dedup por el flag, no por el CUIL).
        records = [_row(name="No identificado", cuil=_VALID_CUIL)]
        result = await apply_import(repo, tid, records)
        # Crea uno NUEVO (no matchea contra el sentinela, que está excluido del índice).
        assert len(result.created_ids) == 1
        assert result.created_ids[0] != sentinel.id

    async def test_brand_collapsed_never_touched_by_import(
        self, db_session: Any, sample_tenant: Any
    ) -> None:
        from app.persistence.repositories.supplier_repository import SupplierRepository

        repo = SupplierRepository(db_session)
        tid = sample_tenant.tenant_id
        collapsed = Supplier(
            tenant_id=tid,
            name="Marca Colapsada",
            cuil=_VALID_CUIL,
            custom_fields={BRAND_COLLAPSED_FLAG_KEY: "true"},
        )
        db_session.add(collapsed)
        await db_session.commit()

        records = [_row(name="Marca Colapsada", cuil=_VALID_CUIL)]
        result = await apply_import(repo, tid, records)
        # No matchea la marca colapsada (excluida del índice) → crea un proveedor nuevo.
        assert len(result.created_ids) == 1
        assert result.created_ids[0] != collapsed.id
