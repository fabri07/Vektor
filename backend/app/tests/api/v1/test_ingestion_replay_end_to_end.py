"""F-H3 — la compuerta: apertura 10 + compra 5 − venta 4 tiene que dar 11.

Los tests unitarios de F-H3 prueban cada pieza contra fixtures que ellos mismos
arman, y eso deja un agujero que ya se pagó en otras fases: la fórmula puede estar
bien mientras el endpoint ignora el modo elegido, aplica una hoja de más o
persiste un saldo distinto del que mostró. Acá el número cruza el archivo real
—un `.xlsx` de tres hojas parseado por el parser de producción—, el confirm, el
apply, la persistencia y la lectura del saldo.

El libro se arma con openpyxl adentro del test en vez de versionar un binario:
así se lee qué contiene sin abrirlo con Excel, y un cambio de clasificación del
parser rompe acá con el header a la vista.

**Qué pasa en cada momento, que no es obvio:**

- Al confirmar, el catálogo deja el stock en 10 (apertura) y la compra suma 5. Las
  compras del import aplican SIEMPRE, sin mirar `inventory_effect` (**V16**): el
  eje gobierna las ventas. Por eso después de confirmar hay 15, no 11.
- La venta descuenta recién en el segundo paso (`/inventory-replay`), que es la
  decisión de F-H3.c: confirmar → revisar → aplicar.
- Con `informational` el saldo se queda en 15 y el preview igual dice 11: el modo
  es de la HOJA DE VENTAS, no del archivo — la apertura y la compra ya lo tocaron.
"""

from __future__ import annotations

import io
import uuid
from typing import Any

import pytest
import pytest_asyncio
from httpx import AsyncClient
from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.application.services.file_parsing import parse_uploaded_content
from app.persistence.models.file import PROCESSING_STATUS_NEEDS_CONFIRMATION, UploadedFile
from app.persistence.models.inventory import InventoryMovement
from app.persistence.models.product import Product
from app.persistence.models.tenant import Tenant
from app.persistence.models.transaction import ExpenseEntry, SaleEntry

_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_PRODUCTO = "Vela aromatica 200g"


@pytest.fixture(autouse=True)
def _sin_broker(mock_score_trigger: Any) -> None:
    """Sin broker en tests, cada confirm/replay/DELETE pagaba ~5s de reintentos
    de kombu (`trigger_score_recalculation.delay`). El servicio traga el error
    (fail-safe): ningún assert dependía del encolado. Con el mock, el mismo
    camino e2e corre en segundos en vez de ~58s."""

_CATALOGO = "sheet:Catalogo"
_COMPRAS = "sheet:Compras"
_VENTAS = "sheet:Ventas"

#: apertura 10 + compra 5. Lo que hay DESPUÉS de confirmar, antes de aplicar nada.
_STOCK_TRAS_CONFIRMAR = 15
#: … − venta 4. Lo que tiene que quedar después del replay.
_STOCK_FINAL = 11


def _libro() -> bytes:
    """Tres hojas: saldo inicial, una compra y una venta del mismo producto."""
    wb = Workbook()
    catalogo = wb.active
    catalogo.title = "Catalogo"
    catalogo.append(["producto", "stock", "precio de venta", "precio de compra"])
    catalogo.append([_PRODUCTO, 10, 2100, 1200])

    compras = wb.create_sheet("Compras")
    compras.append(["fecha", "producto", "cantidad", "total", "proveedor", "forma de pago"])
    compras.append(["2024-03-05", _PRODUCTO, 5, 6000, "Distribuidora Sur", "efectivo"])

    ventas = wb.create_sheet("Ventas")
    ventas.append(["fecha", "producto", "cantidad", "total", "cliente", "forma de pago"])
    ventas.append(["2024-03-10", _PRODUCTO, 4, 8400, "Consumidor final", "efectivo"])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _map(source: str, target: str, context_id: str, entity: str) -> dict[str, Any]:
    return {
        "source_column": source,
        "target_field": target,
        "context_id": context_id,
        "entity_type": entity,
    }


_MAPEOS: list[dict[str, Any]] = [
    _map("producto", "name", _CATALOGO, "product"),
    _map("stock", "stock_units", _CATALOGO, "product"),
    _map("precio de venta", "sale_price_ars", _CATALOGO, "product"),
    _map("precio de compra", "unit_cost_ars", _CATALOGO, "product"),
    _map("fecha", "expense_date", _COMPRAS, "expense"),
    _map("producto", "product_name", _COMPRAS, "expense"),
    _map("cantidad", "quantity", _COMPRAS, "expense"),
    _map("total", "amount", _COMPRAS, "expense"),
    _map("proveedor", "supplier_name", _COMPRAS, "expense"),
    _map("forma de pago", "payment_method", _COMPRAS, "expense"),
    _map("fecha", "transaction_date", _VENTAS, "sale"),
    _map("producto", "product_name", _VENTAS, "sale"),
    _map("cantidad", "quantity", _VENTAS, "sale"),
    _map("total", "amount", _VENTAS, "sale"),
    _map("cliente", "customer_name", _VENTAS, "sale"),
    _map("forma de pago", "payment_method", _VENTAS, "sale"),
]


def _payload(efecto_ventas: str) -> dict[str, Any]:
    return {
        "column_mappings": _MAPEOS,
        "confirmed_fields": {"productos": True, "gastos": True, "ventas": True},
        "context_confirmed": {_CATALOGO: True, _COMPRAS: True, _VENTAS: True},
        # El catálogo es un saldo de apertura, no una compra: no genera COGS ni
        # baja de caja. Es el eje CONTABLE, distinto del de unidades.
        "stock_treatment": {_CATALOGO: "opening_balance", _COMPRAS: "purchase"},
        "inventory_effect": {_VENTAS: efecto_ventas},
    }


@pytest_asyncio.fixture
async def archivo(db_session: AsyncSession, sample_tenant: Tenant) -> UploadedFile:
    """El `.xlsx` pasado por el parser de producción, no un summary a mano."""
    summary = parse_uploaded_content(_libro(), _XLSX_MIME, "libro_completo.xlsx")
    # Si el parser dejara de clasificar así, el resto del test mediría otra cosa.
    assert summary["multi_sheet"] is True
    assert [
        (c["context_id"], c["entity_type"]) for c in summary["mapping_contexts"]
    ] == [(_CATALOGO, "product"), (_COMPRAS, "expense"), (_VENTAS, "sale")]

    record = UploadedFile(
        tenant_id=sample_tenant.tenant_id,
        uploaded_by=None,
        original_filename="libro_completo.xlsx",
        s3_key=f"uploads/test/{uuid.uuid4()}/libro_completo.xlsx",
        content_type=_XLSX_MIME,
        size_bytes=4096,
        purpose="ingestion",
        status="uploaded",
        processing_status=PROCESSING_STATUS_NEEDS_CONFIRMATION,
        parsed_summary_json=summary,
    )
    db_session.add(record)
    await db_session.commit()
    return record


async def _cuantas(db_session: AsyncSession, tenant: Tenant, modelo: Any) -> int:
    """Filas vivas del tenant para ese modelo."""
    result = await db_session.execute(
        select(modelo).where(
            modelo.tenant_id == tenant.tenant_id, modelo.voided_at.is_(None)
        )
        if hasattr(modelo, "voided_at")
        else select(modelo).where(modelo.tenant_id == tenant.tenant_id)
    )
    return len(list(result.scalars().all()))


async def _producto(db_session: AsyncSession, tenant: Tenant) -> Product:
    result = await db_session.execute(
        select(Product).where(Product.tenant_id == tenant.tenant_id)
    )
    productos = list(result.scalars().all())
    assert len(productos) == 1, f"se esperaba 1 producto, hay {len(productos)}"
    await db_session.refresh(productos[0])
    return productos[0]


async def _stock(db_session: AsyncSession, tenant: Tenant) -> int:
    return int((await _producto(db_session, tenant)).stock_units)


async def _movimientos(db_session: AsyncSession, tenant: Tenant) -> list[tuple[str, int]]:
    """`(tipo, cantidad)` de los movimientos vivos, ordenados por tipo.

    Por TIPO y no por fecha a propósito: el saldo de apertura no tiene fecha de
    negocio (es un saldo, no un movimiento fechado) y ordenarlo por
    `occurred_at` lo mandaría al final, detrás de una compra de 2024. Lo que este
    helper afirma es el contenido del ledger, no su cronología.
    """
    result = await db_session.execute(
        select(InventoryMovement).where(
            InventoryMovement.tenant_id == tenant.tenant_id,
            InventoryMovement.voided_at.is_(None),
        )
    )
    movimientos = [(m.movement_type, int(m.qty)) for m in result.scalars().all()]
    return sorted(movimientos)


async def _confirmar(
    client: AsyncClient,
    auth_headers: dict[str, Any],
    archivo: UploadedFile,
    efecto_ventas: str,
) -> dict[str, Any]:
    response = await client.post(
        f"/api/v1/ingestion/files/{archivo.id}/confirm",
        json=_payload(efecto_ventas),
        headers=auth_headers,
    )
    assert response.status_code == 200, response.text
    return dict(response.json())


async def _replay(
    client: AsyncClient,
    auth_headers: dict[str, Any],
    archivo: UploadedFile,
    *,
    dry_run: bool,
) -> dict[str, Any]:
    response = await client.post(
        f"/api/v1/ingestion/files/{archivo.id}/inventory-replay",
        json={"dry_run": dry_run, "context_ids": [_VENTAS]},
        headers=auth_headers,
    )
    assert response.status_code == 200, response.text
    return dict(response.json())


@pytest.mark.asyncio
@pytest.mark.slow
class TestReplayEndToEnd:
    """El ÚNICO e2e del pipeline completo de replay — no borrar ni bajar de capa.

    Los otros invariantes (idempotencia, reversión por borrado, informational)
    viven en capa servicio, donde el mismo assert corre en milisegundos:
    - idempotencia → `test_inventory_replay_service.py::
      test_aplicar_dos_veces_no_descuenta_dos_veces`
    - reversión del replay al borrar → `test_inventory_replay_service.py::
      TestBorrarElArchivoDeshaceElReplay` + ciclo DELETE completo en
      `test_file_deletion_end_to_end.py`
    - informational por hoja (V16) → `test_ingestion_stock_treatment_por_hoja.py::
      test_informational_en_ventas_no_frena_la_apertura_ni_la_compra`
    """

    async def test_apertura_mas_compra_menos_venta_da_once(
        self,
        client: AsyncClient,
        auth_headers: dict[str, Any],
        db_session: AsyncSession,
        sample_tenant: Tenant,
        archivo: UploadedFile,
    ) -> None:
        """El recorrido completo: confirmar → revisar → aplicar → leer el saldo."""
        await _confirmar(client, auth_headers, archivo, "historical_replay")
        # Una fila de cada hoja entró como lo que la hoja dice ser. Se cuenta en la
        # DB y no en la respuesta del confirm: lo que importa es lo que quedó
        # persistido, que es lo que después lee el replay.
        assert await _cuantas(db_session, sample_tenant, SaleEntry) == 1
        assert await _cuantas(db_session, sample_tenant, ExpenseEntry) == 1
        assert await _cuantas(db_session, sample_tenant, Product) == 1
        # El confirm no descuenta la venta: apertura + compra y nada más.
        assert await _stock(db_session, sample_tenant) == _STOCK_TRAS_CONFIRMAR

        # El preview proyecta el saldo final SIN escribirlo.
        preview = await _replay(client, auth_headers, archivo, dry_run=True)
        assert preview["aplicadas"] == 0
        assert [(p["saldo_inicial"], p["saldo_final"]) for p in preview["impacto"]] == [
            (_STOCK_TRAS_CONFIRMAR, _STOCK_FINAL)
        ]
        assert preview["hojas"] == [_VENTAS]
        assert preview["alcance_por_hoja"] is True
        assert await _stock(db_session, sample_tenant) == _STOCK_TRAS_CONFIRMAR

        aplicado = await _replay(client, auth_headers, archivo, dry_run=False)
        assert aplicado["aplicadas"] == 1
        assert aplicado["sin_stock"] == []
        assert await _stock(db_session, sample_tenant) == _STOCK_FINAL

        # El ledger tiene que contar la misma historia que el saldo: +10 de
        # apertura, +5 de la compra y −4 de la venta. Un saldo correcto con un
        # ledger que no lo explica rompe el chequeo de integridad más adelante.
        assert await _movimientos(db_session, sample_tenant) == [
            ("adjustment", 10),
            ("purchase", 5),
            ("sale", -4),
        ]


@pytest.mark.asyncio
@pytest.mark.slow
class TestElApplyExigeElegirHojas:
    """El eje de inventario se declara POR HOJA: escribir sin decir sobre cuáles
    contradice esa declaración.

    Un libro con una hoja de ventas de servicios (`no_inventory`, sus filas no
    mueven unidades) y otra de mercadería descontaba las dos: el panel mandaba
    `context_ids=null` y el servicio interpretaba "todas". El preview sí puede
    correr sobre el archivo entero — es read-only y es la forma en que la
    pantalla descubre qué hojas ofrecer.
    """

    @pytest.mark.parametrize(
        "payload",
        [
            pytest.param({"dry_run": False}, id="test_aplicar_sin_hojas_es_422"),
            # `[]` es "ninguna", no "todas": el default del contrato no puede
            # colarse por una lista vacía.
            pytest.param(
                {"dry_run": False, "context_ids": []},
                id="test_una_lista_vacia_tampoco_alcanza",
            ),
        ],
    )
    async def test_aplicar_sin_elegir_hojas_es_422(
        self,
        client: AsyncClient,
        auth_headers: dict[str, Any],
        archivo: UploadedFile,
        db_session: AsyncSession,
        sample_tenant: Tenant,
        payload: dict[str, Any],
    ) -> None:
        await _confirmar(client, auth_headers, archivo, "historical_replay")

        respuesta = await client.post(
            f"/api/v1/ingestion/files/{archivo.id}/inventory-replay",
            json=payload,
            headers=auth_headers,
        )
        assert respuesta.status_code == 422
        assert "Elegí qué hojas" in respuesta.json()["detail"]

        # Y no escribió nada: el stock sigue como lo dejó el confirm.
        assert await _stock(db_session, sample_tenant) == _STOCK_TRAS_CONFIRMAR

    async def test_el_preview_sigue_viendo_todo_el_archivo(
        self,
        client: AsyncClient,
        auth_headers: dict[str, Any],
        archivo: UploadedFile,
    ) -> None:
        """Sin esto la pantalla no tendría de dónde sacar las hojas para ofrecer."""
        await _confirmar(client, auth_headers, archivo, "historical_replay")

        respuesta = await client.post(
            f"/api/v1/ingestion/files/{archivo.id}/inventory-replay",
            json={"dry_run": True},
            headers=auth_headers,
        )
        assert respuesta.status_code == 200
        assert respuesta.json()["hojas"] == [_VENTAS]
