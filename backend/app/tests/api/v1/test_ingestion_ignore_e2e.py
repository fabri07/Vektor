"""E2 — una columna marcada `ignore` no puede alimentar ningún efecto de negocio.

Por qué e2e y no unitario
-------------------------
El defecto no estaba en una función: estaba en el ORDEN. El importador resolvía
las doce columnas por heurística sobre los headers ANTES de mirar el mapeo, y
recién después el mapeo las pisaba con ``or``. Una columna en `ignore` no entraba
a ``target_to_col`` —``_resolve_target_cols`` la salteaba con un ``continue``—,
así que el ``.get()`` devolvía ``None`` y ganaba la heurística. Tampoco entraba a
``_reservadas`` ni al ``skip`` de ``_row_val``, de modo que quedaba libre incluso
para ser reasignada al monto.

Un test sobre ``_resolve_target_cols`` no ve nada de eso: la función "cumple"
—devuelve la columna en ningún campo— y el bug vive en lo que pasa después. Por
eso acá el `.xlsx` lo arma openpyxl, lo parsea el parser de producción, el confirm
entra por HTTP y lo que se afirma es la ENTIDAD PERSISTIDA.

Cada aserción elige una columna que la heurística tomaría igual: `proveedor`
matchea ``_PROVEEDOR_COLS``, `cantidad` matchea ``_CANTIDAD_COLS``, `precio de
compra` matchea ``_COSTO_COLS``. Si el `ignore` no se respetara, el valor del
archivo aparecería en la fila guardada.
"""

from __future__ import annotations

import io
import uuid
from typing import Any

import pytest
import pytest_asyncio
from httpx import AsyncClient
from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.application.services.file_parsing import parse_uploaded_content
from app.persistence.models.file import PROCESSING_STATUS_NEEDS_CONFIRMATION, UploadedFile
from app.persistence.models.product import Product
from app.persistence.models.supplier import Supplier
from app.persistence.models.tenant import Tenant
from app.persistence.models.transaction import SaleEntry
from app.persistence.models.unclassified_record import UnclassifiedRecord

_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_PRODUCTO = "Vela aromatica 200g"
_PROVEEDOR = "Distribuidora Sur"
_CATALOGO = "sheet:Catalogo"
_COMPRAS = "sheet:Compras"
_VENTAS = "sheet:Ventas"


@pytest.fixture(autouse=True)
def _sin_broker(mock_score_trigger: Any) -> None:
    """Sin broker, cada confirm paga ~5s de reintentos de kombu (fail-safe)."""


def _libro_multihoja() -> bytes:
    wb = Workbook()
    catalogo = wb.active
    catalogo.title = "Catalogo"
    catalogo.append(["producto", "stock", "precio de venta", "precio de compra"])
    catalogo.append([_PRODUCTO, 10, 2100, 1200])

    compras = wb.create_sheet("Compras")
    compras.append(["fecha", "producto", "cantidad", "total", "proveedor", "forma de pago"])
    compras.append(["2024-03-05", _PRODUCTO, 5, 6000, _PROVEEDOR, "efectivo"])

    ventas = wb.create_sheet("Ventas")
    ventas.append(["fecha", "producto", "cantidad", "total", "cliente", "forma de pago"])
    ventas.append(["2024-03-10", _PRODUCTO, 4, 8400, "Consumidor final", "efectivo"])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _tabla_de_ventas() -> bytes:
    """Una sola hoja: el camino PLANO, que resuelve columnas distinto del multihoja."""
    wb = Workbook()
    hoja = wb.active
    hoja.title = "Ventas"
    hoja.append(["fecha", "producto", "cantidad", "total", "cliente", "forma de pago"])
    hoja.append(["2024-03-10", _PRODUCTO, 4, 8400, "Consumidor final", "efectivo"])
    hoja.append(["2024-03-11", _PRODUCTO, 3, 6300, "Consumidor final", "efectivo"])
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _map(source: str, target: str, context_id: str | None, entity: str | None) -> dict[str, Any]:
    fila: dict[str, Any] = {"source_column": source, "target_field": target}
    if context_id is not None:
        fila["context_id"] = context_id
        fila["entity_type"] = entity
    return fila


async def _subir(
    db_session: AsyncSession, tenant: Tenant, contenido: bytes, nombre: str
) -> UploadedFile:
    summary = parse_uploaded_content(contenido, _XLSX_MIME, nombre)
    record = UploadedFile(
        tenant_id=tenant.tenant_id,
        uploaded_by=None,
        original_filename=nombre,
        s3_key=f"uploads/test/{uuid.uuid4()}/{nombre}",
        content_type=_XLSX_MIME,
        size_bytes=4096,
        purpose="ingestion",
        status="uploaded",
        processing_status=PROCESSING_STATUS_NEEDS_CONFIRMATION,
        parsed_summary_json=summary,
    )
    db_session.add(record)
    await db_session.commit()
    return record


@pytest_asyncio.fixture
async def archivo_multihoja(db_session: AsyncSession, sample_tenant: Tenant) -> UploadedFile:
    record = await _subir(db_session, sample_tenant, _libro_multihoja(), "libro.xlsx")
    summary = record.parsed_summary_json
    # Si el parser dejara de clasificar así, el test mediría otra cosa.
    assert summary["multi_sheet"] is True
    return record


@pytest_asyncio.fixture
async def archivo_plano(db_session: AsyncSession, sample_tenant: Tenant) -> UploadedFile:
    record = await _subir(db_session, sample_tenant, _tabla_de_ventas(), "ventas.xlsx")
    assert not record.parsed_summary_json.get("multi_sheet")
    return record


async def _ventas(db_session: AsyncSession, tenant: Tenant) -> list[SaleEntry]:
    result = await db_session.execute(
        select(SaleEntry).where(
            SaleEntry.tenant_id == tenant.tenant_id, SaleEntry.voided_at.is_(None)
        )
    )
    return list(result.scalars().all())


async def test_multihoja_ignora_proveedor_cantidad_y_costo(
    client: AsyncClient,
    auth_headers: dict[str, str],
    db_session: AsyncSession,
    sample_tenant: Tenant,
    archivo_multihoja: UploadedFile,
) -> None:
    """Tres columnas ignoradas, tres heurísticas que las habrían recuperado."""
    mapeos = [
        _map("producto", "name", _CATALOGO, "product"),
        _map("stock", "stock_units", _CATALOGO, "product"),
        _map("precio de venta", "sale_price_ars", _CATALOGO, "product"),
        # `_COSTO_COLS` matchea "precio de compra": sin el fix, el costo entraba igual.
        _map("precio de compra", "ignore", _CATALOGO, "product"),
        _map("fecha", "expense_date", _COMPRAS, "expense"),
        _map("producto", "product_name", _COMPRAS, "expense"),
        _map("cantidad", "quantity", _COMPRAS, "expense"),
        _map("total", "amount", _COMPRAS, "expense"),
        # `_PROVEEDOR_COLS` matchea "proveedor".
        _map("proveedor", "ignore", _COMPRAS, "expense"),
        _map("forma de pago", "payment_method", _COMPRAS, "expense"),
        _map("fecha", "transaction_date", _VENTAS, "sale"),
        _map("producto", "product_name", _VENTAS, "sale"),
        # `_CANTIDAD_COLS` matchea "cantidad".
        _map("cantidad", "ignore", _VENTAS, "sale"),
        _map("total", "amount", _VENTAS, "sale"),
        _map("cliente", "customer_name", _VENTAS, "sale"),
        _map("forma de pago", "payment_method", _VENTAS, "sale"),
    ]
    resp = await client.post(
        f"/api/v1/ingestion/files/{archivo_multihoja.id}/confirm",
        json={
            "column_mappings": mapeos,
            "confirmed_fields": {"productos": True, "gastos": True, "ventas": True},
            "context_confirmed": {_CATALOGO: True, _COMPRAS: True, _VENTAS: True},
            "stock_treatment": {_CATALOGO: "opening_balance", _COMPRAS: "purchase"},
        },
        headers=auth_headers,
    )
    assert resp.status_code == 200, resp.text

    # 1. El proveedor del archivo NO se creó: la compra fue al centinela.
    proveedores = list(
        (
            await db_session.execute(
                select(Supplier).where(Supplier.tenant_id == sample_tenant.tenant_id)
            )
        ).scalars()
    )
    assert _PROVEEDOR not in [s.name for s in proveedores], (
        f"la columna ignorada creó el proveedor: {[s.name for s in proveedores]}"
    )

    # 2. La venta no tomó la cantidad del archivo (quantity es NOT NULL default 1).
    ventas = await _ventas(db_session, sample_tenant)
    assert len(ventas) == 1
    assert ventas[0].quantity != 4, "la cantidad ignorada entró igual a la venta"
    assert ventas[0].amount == 8400, "el monto mapeado sí tiene que entrar"

    # 3. El costo del catálogo no salió de la columna ignorada.
    producto = (
        await db_session.execute(
            select(Product).where(Product.tenant_id == sample_tenant.tenant_id)
        )
    ).scalars().first()
    assert producto is not None
    assert producto.unit_cost_ars != 1200, "el costo ignorado entró por la heurística"


async def test_plano_ignora_la_cantidad(
    client: AsyncClient,
    auth_headers: dict[str, str],
    db_session: AsyncSession,
    sample_tenant: Tenant,
    archivo_plano: UploadedFile,
) -> None:
    """El camino de UNA tabla resuelve columnas por su cuenta y necesita su propio caso.

    Sin `context_id` los mapeos son "planos" y el importador entra por
    ``_insert_confirmed_data_impl``, donde la heurística corre sobre
    ``rows[0].keys()`` — otra rama, mismo defecto.
    """
    resp = await client.post(
        f"/api/v1/ingestion/files/{archivo_plano.id}/confirm",
        json={
            "column_mappings": [
                _map("fecha", "transaction_date", None, None),
                _map("producto", "product_name", None, None),
                _map("cantidad", "ignore", None, None),
                _map("total", "amount", None, None),
                _map("cliente", "customer_name", None, None),
                _map("forma de pago", "payment_method", None, None),
            ],
            "confirmed_fields": {"ventas": True},
        },
        headers=auth_headers,
    )
    assert resp.status_code == 200, resp.text

    ventas = await _ventas(db_session, sample_tenant)
    assert len(ventas) == 2, f"se esperaban las dos filas, hay {len(ventas)}"
    assert {v.quantity for v in ventas} == {1}, (
        f"la cantidad ignorada entró igual: {[v.quantity for v in ventas]}"
    )
    assert sorted(v.amount for v in ventas) == [6300, 8400]


def _tabla_con_fila_sin_monto() -> bytes:
    """Segunda fila sin total: cae a "Otros" con motivo, no se pierde."""
    wb = Workbook()
    hoja = wb.active
    hoja.title = "Ventas"
    hoja.append(["fecha", "producto", "cantidad", "total", "cliente", "forma de pago"])
    hoja.append(["2024-03-10", _PRODUCTO, 4, 8400, "Consumidor final", "efectivo"])
    hoja.append(["2024-03-11", _PRODUCTO, 3, None, "Consumidor final", "efectivo"])
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


async def test_la_bandeja_otros_hereda_la_decision(
    client: AsyncClient,
    auth_headers: dict[str, str],
    db_session: AsyncSession,
    sample_tenant: Tenant,
) -> None:
    """Una fila derivada a "Otros" no puede llevarse la columna ignorada adentro.

    ``import_unclassified_records`` reimporta desde la bandeja leyendo el
    ``row_data`` **100% por keyword**, sin consultar ningún mapeo: es la puerta
    de atrás por la que la decisión volvería a perderse. Como el saneo se aplica
    sobre las filas ANTES de que el importador las mire, lo que se captura ya
    viene sin esa columna y la bandeja hereda la decisión por construcción.

    Límite conocido: las filas capturadas por importaciones ANTERIORES a este
    cambio sí tienen la columna guardada. No es recuperable sin las decisiones
    de ese import, que no se persistían.
    """
    record = await _subir(
        db_session, sample_tenant, _tabla_con_fila_sin_monto(), "ventas_parcial.xlsx"
    )
    resp = await client.post(
        f"/api/v1/ingestion/files/{record.id}/confirm",
        json={
            "column_mappings": [
                _map("fecha", "transaction_date", None, None),
                _map("producto", "product_name", None, None),
                _map("cantidad", "ignore", None, None),
                _map("total", "amount", None, None),
                _map("cliente", "customer_name", None, None),
                _map("forma de pago", "payment_method", None, None),
            ],
            "confirmed_fields": {"ventas": True},
        },
        headers=auth_headers,
    )
    assert resp.status_code == 200, resp.text

    pendientes = list(
        (
            await db_session.execute(
                select(UnclassifiedRecord).where(
                    UnclassifiedRecord.tenant_id == sample_tenant.tenant_id
                )
            )
        ).scalars()
    )
    assert len(pendientes) == 1, f"se esperaba 1 fila en Otros, hay {len(pendientes)}"
    guardada = pendientes[0].row_data
    assert "cantidad" not in guardada, (
        f"la columna ignorada viajó a Otros: {sorted(guardada)}"
    )
    # Y lo que el usuario SÍ mapeó sigue estando, para poder corregir la fila.
    assert guardada.get("producto") == _PRODUCTO
